"""Seed Novza departments + employees + 9-18 shift from an Excel file.

Reads the "Апрель" sheet of the Novza табел Excel, groups employees by
department header rows (Столярка / Малярка / ОФИС / Упакофка / Охрана),
creates a 9:00-18:00 shift template, then provisions:

  - departments (one per Excel section)
  - users with login = ``novza_<first_name>``  (numeric suffix on collision)
  - employees linked to user + dept + branch + shift template
  - hire_date forced to a single calendar day (default 2026-05-04)

The script is idempotent on (company_id, full_name, department_id) — a
re-run won't duplicate rows. Credentials of newly created users are
printed to stdout AND written to a CSV next to the Excel for safe-keeping.

USAGE (inside the api container):
    python -m scripts.seed_novza /tmp/novza.xlsx --company novza

Optional flags:
    --branch "Asosiy filial"
    --hire-date 2026-05-04
    --prefix novza_
    --dry-run                # parse and print what WOULD be created
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import re
import sys
from datetime import date, time
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.permissions import Role
from app.core.security import hash_password
from app.core.tenant import install_tenant_listener, set_current_tenant
from app.database import AsyncSessionLocal
from app.models.branch import Branch
from app.models.company import Company
from app.models.department import Department
from app.models.employee import Employee, SalaryType, WorkType
from app.models.shift import ShiftTemplate, ShiftType
from app.models.user import User, UserStatus

DEPARTMENTS = {"Столярка", "Малярка", "ОФИС", "Упакофка", "Охрана"}
HEADER_TOKENS = {"ф.и.о", "ф.и.о.", "должность", "fio"}

# Cyrillic + Uzbek-specific → Latin. Lossy on purpose so a name with
# diacritics still produces a valid Latin login slug.
CYR_TO_LATIN: dict[str, str] = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e",
    "ё": "yo", "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k",
    "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
    "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "",
    "э": "e", "ю": "yu", "я": "ya",
    # Uzbek-Cyrillic specials
    "ў": "o", "ғ": "g", "қ": "q", "ҳ": "h", "ҷ": "j", "ӣ": "i",
}


def translit(text: str) -> str:
    out: list[str] = []
    for ch in text.lower():
        out.append(CYR_TO_LATIN.get(ch, ch))
    return re.sub(r"[^a-z0-9]+", "", "".join(out))


def first_name(full: str) -> str:
    """Pick the second word as the given name. Common Uzbek convention is
    Surname Firstname [Middlename]."""
    parts = [p for p in full.strip().split() if p]
    if len(parts) >= 2:
        return translit(parts[1])
    return translit(parts[0]) if parts else ""


def is_index(s: str) -> bool:
    return bool(re.match(r"^\d+$", s.strip()))


def parse_excel(path: Path) -> list[tuple[str, str, str | None]]:
    """Yield (department, full_name, position) for every employee row.

    Row shape: column A may be numeric index OR — for section headers —
    the dept name. Column B is the F.I.O. Column C is the position
    (Должность). A handful of legacy rows have two names spread over
    A + B because of merged cells; both are emitted.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["Апрель"]
    current_dept: str | None = None
    out: list[tuple[str, str, str | None]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        a = (str(row[0]).strip() if len(row) > 0 and row[0] is not None else "")
        b = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        c = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else "")

        if a in DEPARTMENTS or b in DEPARTMENTS:
            current_dept = a if a in DEPARTMENTS else b
            continue
        if not current_dept or not (a or b):
            continue
        if a.lower() in HEADER_TOKENS or b.lower() in HEADER_TOKENS:
            continue

        names: list[str] = []
        if b and not is_index(b):
            names.append(b)
        if a and not is_index(a) and a != b and a not in DEPARTMENTS:
            names.append(a)
        for nm in names:
            out.append((current_dept, nm, c or None))
    return out


async def get_or_create_dept(
    db: AsyncSession, *, company_id, branch_id, name: str
) -> Department:
    existing = (
        await db.execute(
            select(Department).where(
                Department.company_id == company_id,
                Department.branch_id == branch_id,
                Department.name == name,
            )
        )
    ).scalar_one_or_none()
    if existing:
        return existing
    d = Department(
        company_id=company_id,
        branch_id=branch_id,
        name=name,
        is_active=True,
    )
    db.add(d)
    await db.flush()
    return d


async def get_or_create_shift(
    db: AsyncSession, *, company_id
) -> ShiftTemplate:
    name = "9:00–18:00"
    existing = (
        await db.execute(
            select(ShiftTemplate).where(
                ShiftTemplate.company_id == company_id,
                ShiftTemplate.name == name,
            )
        )
    ).scalar_one_or_none()
    if existing:
        return existing
    s = ShiftTemplate(
        company_id=company_id,
        name=name,
        type=ShiftType.FIXED.value,
        start_time=time(9, 0),
        end_time=time(18, 0),
        break_minutes=60,
        expected_hours=8.0,
        allow_overtime=True,
        is_active=True,
    )
    db.add(s)
    await db.flush()
    return s


async def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Seed Novza employees from Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("excel", type=Path, help="Path to the Novza табел .xlsx")
    parser.add_argument("--company", default="novza", help="Company slug (default: novza)")
    parser.add_argument("--branch", default="Asosiy filial", help="Branch name")
    parser.add_argument("--hire-date", default="2026-05-04", help="YYYY-MM-DD")
    parser.add_argument("--prefix", default="novza_", help="Login/password prefix")
    parser.add_argument("--dry-run", action="store_true", help="Don't write to DB")
    args = parser.parse_args(argv)

    if not args.excel.exists():
        print(f"❌ Excel not found: {args.excel}", file=sys.stderr)
        return 1

    hire_date = date.fromisoformat(args.hire_date)
    rows = parse_excel(args.excel)
    dept_set = {r[0] for r in rows}
    print(f"📄 Parsed {len(rows)} employees across {len(dept_set)} departments")
    for d in DEPARTMENTS:
        n = sum(1 for r in rows if r[0] == d)
        if n:
            print(f"     • {d}: {n}")

    if args.dry_run:
        print("\n--- DRY RUN — first 10 rows ---")
        for d, n, p in rows[:10]:
            print(f"  [{d:10s}] {n!r:40s}  {p or ''}")
        return 0

    install_tenant_listener()

    async with AsyncSessionLocal() as db:
        company = (
            await db.execute(
                select(Company)
                .where(Company.slug == args.company)
                .execution_options(skip_tenant_filter=True)
            )
        ).scalar_one_or_none()
        if not company:
            print(
                f"❌ Company '{args.company}' not found. "
                f"Create it first via /owner/companies/new (slug={args.company}).",
                file=sys.stderr,
            )
            return 2
        print(f"🏢 Company: {company.name} (id={company.id})")
        set_current_tenant(company.id)

        branch = (
            await db.execute(
                select(Branch).where(
                    Branch.company_id == company.id,
                    Branch.name == args.branch,
                )
            )
        ).scalar_one_or_none()
        if not branch:
            branch = Branch(
                company_id=company.id,
                name=args.branch,
                is_active=True,
                geofence_radius_m=150,
            )
            db.add(branch)
            await db.flush()
        print(f"🏬 Branch:  {branch.name} (id={branch.id})")

        shift = await get_or_create_shift(db, company_id=company.id)
        print(f"⏰ Shift:   {shift.name}")

        existing_usernames: set[str] = set(
            (
                await db.execute(
                    select(User.username)
                    .where(User.company_id == company.id)
                    .execution_options(skip_tenant_filter=True)
                )
            ).scalars().all()
        )
        existing_codes = (
            await db.execute(
                select(Employee.employee_code).where(
                    Employee.company_id == company.id
                )
            )
        ).scalars().all()
        used_code_nums = {
            int(m.group(1))
            for code in existing_codes
            if (m := re.match(r"^NV(\d+)$", code or ""))
        }
        next_code = 1

        def next_emp_code() -> str:
            nonlocal next_code
            while next_code in used_code_nums:
                next_code += 1
            used_code_nums.add(next_code)
            code = f"NV{next_code:03d}"
            next_code += 1
            return code

        depts: dict[str, Department] = {}
        created: list[tuple[str, str, str, str, str]] = []  # (dept, name, position, login, password)
        skipped: list[str] = []

        for dept_name, full_name, position in rows:
            if dept_name not in depts:
                depts[dept_name] = await get_or_create_dept(
                    db,
                    company_id=company.id,
                    branch_id=branch.id,
                    name=dept_name,
                )
            dept = depts[dept_name]

            existing_emp = (
                await db.execute(
                    select(Employee).where(
                        Employee.company_id == company.id,
                        Employee.full_name == full_name,
                        Employee.department_id == dept.id,
                    )
                )
            ).scalar_one_or_none()
            if existing_emp:
                skipped.append(f"already exists in {dept_name}: {full_name}")
                continue

            ism = first_name(full_name)
            if not ism:
                skipped.append(f"empty first name: {full_name!r}")
                continue
            base = f"{args.prefix}{ism}"
            login = base
            i = 2
            while login in existing_usernames:
                login = f"{base}_{i}"
                i += 1
            existing_usernames.add(login)
            password = login

            user = User(
                company_id=company.id,
                username=login,
                email=None,
                password_hash=hash_password(password),
                role=Role.EMPLOYEE,
                status=UserStatus.ACTIVE,
                full_name=full_name,
                language="uz",
                is_active=True,
            )
            db.add(user)
            await db.flush()

            emp = Employee(
                company_id=company.id,
                branch_id=branch.id,
                department_id=dept.id,
                user_id=user.id,
                employee_code=next_emp_code(),
                full_name=full_name,
                position=position,
                hire_date=hire_date,
                work_type=WorkType.FIXED_SHIFT.value,
                shift_template_id=shift.id,
                salary_type=SalaryType.MONTHLY.value,
                is_active=True,
            )
            db.add(emp)
            await db.flush()

            created.append((dept_name, full_name, position or "", login, password))

        await db.commit()

        # Persist credentials next to the Excel as a CSV.
        out_path = args.excel.with_name("novza_credentials.csv")
        with out_path.open("w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["department", "full_name", "position", "login", "password"])
            for row in created:
                w.writerow(row)

        print()
        print(f"✓ {len(created)} employees created")
        if skipped:
            print(f"⊘ {len(skipped)} skipped:")
            for s in skipped[:10]:
                print(f"    {s}")
            if len(skipped) > 10:
                print(f"    … and {len(skipped) - 10} more")
        print()
        print(f"💾 Credentials saved → {out_path}")
        print()
        print("FIRST 10 LOGINS (full list in CSV):")
        for d, fn, _pos, lg, pw in created[:10]:
            print(f"   {d:10s}  {fn:35s}  {lg:25s}  {pw}")

    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
