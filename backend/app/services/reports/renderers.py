"""PDF + XLSX renderers for report jobs.

The CSV path stays in ``report_service.py`` (the queue's most-used
format). PDF and XLSX live here because they pull in heavier deps
(WeasyPrint for HTML→PDF, openpyxl for spreadsheet), and we want to
keep the CSV path importable on a system without those libs.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from jinja2 import Environment, FileSystemLoader, select_autoescape

from app.models.company import Company
from app.models.report import ReportType
from app.models.user import User

logger = logging.getLogger(__name__)


_TEMPLATE_DIR = Path(__file__).parent / "templates"

_jinja = Environment(
    loader=FileSystemLoader(str(_TEMPLATE_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
    trim_blocks=True,
    lstrip_blocks=True,
)


# Per-report metadata: which columns are numeric (right-aligned in PDF,
# numeric format in XLSX), which cells should render as status pills, and
# how to compute the optional "totals" row.
_NUMERIC_BY_TYPE: dict[str, set[int]] = {
    ReportType.ATTENDANCE_DAILY.value: {7, 8, 9},   # Worked, Late, OT
    ReportType.ATTENDANCE_MONTHLY.value: {5, 6, 7, 8, 9, 10, 11},
    ReportType.SALARY_REGISTER.value: {5, 6, 7, 8, 9, 10, 11, 12},
    ReportType.EMPLOYEE_ROSTER.value: {9, 10, 11},  # Base, Daily, Hourly
    ReportType.KPI_SUMMARY.value: {5, 6, 7, 8, 9, 10},  # counts + reward + penalty
    ReportType.LEAVE_BALANCE.value: {6, 7, 8},  # cap, used, remaining
    ReportType.BONUS_DEDUCTION_REGISTER.value: {6},  # amount
    ReportType.LATE_ABSENCE_TREND.value: {4, 5, 6, 7, 8},
}


_STATUS_VALUES = {
    "PRESENT", "LATE", "ABSENT", "ON_LEAVE", "REST_DAY", "NOT_SCHEDULED",
    "DRAFT", "FINALIZED", "APPROVED", "PARTIALLY_PAID", "PAID",
    "PENDING", "RUNNING", "READY", "FAILED",
}


def _compute_totals(report_type: str, rows: list[list[str]]) -> list[str] | None:
    """Sum the numeric columns for the report types where a TOTALS row
    actually makes sense. Attendance daily is per-day so totalling rows
    across employees doesn't help — skip it. Roster is per-employee
    static info — also skip."""
    if not rows:
        return None

    if report_type == ReportType.ATTENDANCE_MONTHLY.value:
        # Sum cols 5..11 (days, hours, late, OT, absent, leave, rest)
        total = ["", "", "", "", "TOTAL"]
        for col in range(5, 12):
            if col == 6:
                # "Total hours" is "h:mm" — sum minutes.
                mins = 0
                for r in rows:
                    if len(r) > col and ":" in (r[col] or ""):
                        h, m = r[col].split(":")
                        mins += int(h) * 60 + int(m)
                total.append(f"{mins // 60}:{str(mins % 60).zfill(2)}")
            else:
                s = 0
                for r in rows:
                    try:
                        s += int(r[col])
                    except (ValueError, IndexError):
                        pass
                total.append(str(s))
        return total

    if report_type == ReportType.SALARY_REGISTER.value:
        total = ["", "", "", "", "TOTAL"]
        # Sum cols 5..12 (base, OT, bonuses, KPI, deductions, total, paid, pending)
        for col in range(5, 13):
            s = 0
            for r in rows:
                try:
                    s += int(r[col])
                except (ValueError, IndexError):
                    pass
            total.append(str(s))
        total.append("")  # status column
        return total

    if report_type == ReportType.LATE_ABSENCE_TREND.value:
        # Sum cols 4..8 (late count, late minutes, avg, absent, worst). Avg
        # is left blank because mean-of-means isn't meaningful.
        total = ["", "", "", "TOTAL"]
        for col, summable in [(4, True), (5, True), (6, False), (7, True), (8, False)]:
            if not summable:
                total.append("")
                continue
            s = 0
            for r in rows:
                try:
                    s += int(r[col])
                except (ValueError, IndexError):
                    pass
            total.append(str(s))
        return total

    return None


def render_pdf(
    *,
    report_type: str,
    headers: list[str],
    rows: list[list[str]],
    company: Company,
    requested_by: User | None,
    subtitle: str,
    branch_name: str | None = None,
    locale: str = "uz",
) -> bytes:
    """Render a CSV-shaped table into a paginated PDF via WeasyPrint.

    Returns the PDF bytes ready for upload. WeasyPrint is heavy (cairo
    + pango fonts), so we lazy-import it to keep the API process light
    when no one's running PDFs.
    """
    from weasyprint import HTML

    base_css = (_TEMPLATE_DIR / "_base.css").read_text(encoding="utf-8")
    title_map = {
        ReportType.ATTENDANCE_DAILY.value: {
            "uz": "Davomat — kunlik",
            "ru": "Посещаемость — ежедневная",
            "en": "Attendance — daily",
        },
        ReportType.ATTENDANCE_MONTHLY.value: {
            "uz": "Davomat — oylik",
            "ru": "Посещаемость — ежемесячная",
            "en": "Attendance — monthly",
        },
        ReportType.SALARY_REGISTER.value: {
            "uz": "Oylik registri",
            "ru": "Реестр зарплаты",
            "en": "Salary register",
        },
        ReportType.EMPLOYEE_ROSTER.value: {
            "uz": "Xodimlar ro'yxati",
            "ru": "Список сотрудников",
            "en": "Employee roster",
        },
        ReportType.KPI_SUMMARY.value: {
            "uz": "KPI xulasa",
            "ru": "Сводка по KPI",
            "en": "KPI summary",
        },
        ReportType.LEAVE_BALANCE.value: {
            "uz": "Ta'til balansi",
            "ru": "Баланс отпусков",
            "en": "Leave balance",
        },
        ReportType.BONUS_DEDUCTION_REGISTER.value: {
            "uz": "Bonus va jarima jurnali",
            "ru": "Журнал бонусов и удержаний",
            "en": "Bonus & deduction register",
        },
        ReportType.LATE_ABSENCE_TREND.value: {
            "uz": "Kech kelish va kelmaslik",
            "ru": "Опоздания и прогулы",
            "en": "Late & absence trend",
        },
    }
    label_map = {
        "generated_at": {
            "uz": "Yaratilgan",
            "ru": "Сформировано",
            "en": "Generated",
        },
        "requested_by": {"uz": "So'rovchi", "ru": "Запросил", "en": "Requested by"},
        "branch": {"uz": "Filial", "ru": "Филиал", "en": "Branch"},
        "prepared_by": {"uz": "Tayyorladi", "ru": "Подготовил", "en": "Prepared by"},
        "approved_by": {"uz": "Tasdiqladi", "ru": "Утвердил", "en": "Approved by"},
    }

    locale = locale if locale in ("uz", "ru", "en") else "uz"

    company_name = company.name if company else ""
    company_initial = (company_name or "W")[:1].upper()

    totals = _compute_totals(report_type, rows)

    rendered = _jinja.get_template("report.html.j2").render(
        base_css=base_css,
        report_title=title_map[report_type][locale],
        subtitle=subtitle,
        company_name=company_name,
        company_initial=company_initial,
        generated_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        generated_at_label=label_map["generated_at"][locale],
        requested_by=(
            requested_by.full_name or requested_by.username if requested_by else None
        ),
        requested_by_label=label_map["requested_by"][locale],
        branch_name=branch_name,
        branch_label=label_map["branch"][locale],
        sig_prepared_by_label=label_map["prepared_by"][locale],
        sig_approved_by_label=label_map["approved_by"][locale],
        include_signatures=report_type
        in (
            ReportType.SALARY_REGISTER.value,
            ReportType.ATTENDANCE_MONTHLY.value,
        ),
        headers=headers,
        rows=rows,
        totals=totals,
        numeric_columns=_NUMERIC_BY_TYPE.get(report_type, set()),
        status_pill_values=_STATUS_VALUES,
    )
    return HTML(string=rendered).write_pdf()


def render_xlsx(
    *,
    report_type: str,
    headers: list[str],
    rows: list[list[str]],
    company: Company,
) -> bytes:
    """Render to XLSX via openpyxl. Numeric columns get number format,
    headers get the brand colour, totals row uses a top border."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    totals_font = Font(bold=True)
    totals_fill = PatternFill("solid", fgColor="F1F5F9")
    totals_border = Border(top=Side(style="medium", color="4F46E5"))

    # Title bar
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers))
    )
    title_cell = ws.cell(row=1, column=1, value=company.name if company else "")
    title_cell.font = Font(bold=True, size=14)

    # Header row at row 3
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")

    numeric_cols = _NUMERIC_BY_TYPE.get(report_type, set())

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, cell in enumerate(row, start=1):
            value: Any = cell
            if (c_idx - 1) in numeric_cols:
                # Try to coerce to number; "h:mm" stays a string.
                try:
                    value = int(cell) if cell != "" else 0
                except (ValueError, TypeError):
                    value = cell
            ws.cell(row=r_idx, column=c_idx, value=value)
            if (c_idx - 1) in numeric_cols and isinstance(value, int):
                ws.cell(row=r_idx, column=c_idx).number_format = "#,##0"

    totals = _compute_totals(report_type, rows)
    if totals:
        totals_row = len(rows) + 4
        for c_idx, cell in enumerate(totals, start=1):
            value: Any = cell
            if (c_idx - 1) in numeric_cols:
                try:
                    value = int(cell) if cell != "" else 0
                except (ValueError, TypeError):
                    value = cell
            wcell = ws.cell(row=totals_row, column=c_idx, value=value)
            wcell.font = totals_font
            wcell.fill = totals_fill
            wcell.border = totals_border

    # Auto-fit-ish column widths (capped to keep the file small).
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in rows[:200]:
            if col_idx - 1 < len(row):
                v = row[col_idx - 1]
                if v and len(str(v)) > max_len:
                    max_len = min(40, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(8, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["render_pdf", "render_xlsx"]
